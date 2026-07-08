import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import numbers
from src.immo_data import ImmoData

def create_workbook():
    # create a new workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    workbook.active.title = 'Houses'
    workbook.create_sheet(title="Land")

    # add headers
    for worksheet in workbook.worksheets:
        worksheet['A1'] = 'Type'
        worksheet['B1'] = 'Title'
        worksheet['C1'] = 'Location'
        worksheet['D1'] = 'Price [€]'
        worksheet['E1'] = 'Living Area [m2]'
        worksheet['F1'] = 'Land Area [m2]'
        worksheet['G1'] = 'Ratio [€ / m2]'
        worksheet['H1'] = 'Distance [km]'
        worksheet['I1'] = 'Rating'
        worksheet['J1'] = 'Link'

        for cell in worksheet[1]:
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type="solid")

    return workbook

def write_listings(house_listings: list[ImmoData], land_listings: list[ImmoData]):
    workbook = create_workbook()
    worksheet = workbook.active
    _write_to_worksheet(worksheet, house_listings)
    worksheet = workbook.worksheets[1]
    _write_to_worksheet(worksheet, land_listings)

    workbook.save('immo-results.xlsx')

def _write_to_worksheet(worksheet, listings):
    worksheet.column_dimensions['B'].width = 80
    worksheet.column_dimensions['C'].width = 25
    worksheet.column_dimensions['D'].width = 20
    worksheet.column_dimensions['J'].width = 80
    all_ratings_sorted = sorted([l.rating for l in listings if l.rating is not None], reverse=True)
    for row, listing in enumerate(listings, start=2):
        worksheet.cell(row=row, column=1, value=listing.type.name)
        worksheet.cell(row=row, column=2, value=listing.title)
        worksheet.cell(row=row, column=3, value=listing.location)
        worksheet.cell(row=row, column=4, value=listing.price).number_format=numbers.FORMAT_CURRENCY_EUR_SIMPLE
        worksheet.cell(row=row, column=5, value=listing.living_area)
        worksheet.cell(row=row, column=6, value=listing.land_area)
        worksheet.cell(row=row, column=7, value=listing.ratio)
        worksheet.cell(row=row, column=8, value=listing.distance)
        worksheet.cell(row=row, column=9, value=listing.rating).fill = _get_cell_style(listing.type, listing.rating, all_ratings_sorted)
        worksheet.cell(row=row, column=10, value=listing.link)

def _get_cell_style(type, value, all_ratings):
    green = '008000'
    yellow = 'FFFF00'
    red = 'FF0000'
    color = 'FFFFF0'
    if value is None or value == 0:
        return PatternFill(start_color=color, end_color=color, fill_type='solid')
    limits = type.get_limits()
    index = all_ratings.index(value)
    if index < int(len(all_ratings) * limits[0]):
        color = green
    elif index < int(len(all_ratings) * limits[1]):
        color = yellow
    else:
        color = red

    return PatternFill(start_color=color, end_color=color, fill_type='solid')
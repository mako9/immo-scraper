import os
import openpyxl
from openpyxl.styles import PatternFill
from dotenv import load_dotenv

from src.immo_platform import ImmoPlatform
from src.immowelt import get_immowelt_results
from src.immonet import get_immonet_results
from src.immoscout import get_immoscout_results
from src.rating import calculate_rating

# Load the environment variables from the .env file
load_dotenv()

def _create_workbook():
    # create a new workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    workbook.active.title = 'Houses'
    workbook.create_sheet(title="Land")

    # add headers
    for worksheet in workbook.worksheets:
        worksheet['A1'] = 'Type'
        worksheet['B1'] = 'Title'
        worksheet['C1'] = 'Price [€]'
        worksheet['D1'] = 'Living Area [m2]'
        worksheet['E1'] = 'Land Area [m2]'
        worksheet['F1'] = 'Ratio [€ / m2]'
        worksheet['G1'] = 'Rating'
        worksheet['H1'] = 'Link'

    return workbook


def _get_cell_style(type, value):
    green = '008000'
    yellow = 'FFFF00'
    red = 'FF0000'
    color = 'FFFFF0'
    limits = type.get_limits()
    if value is None:
        return PatternFill(start_color=color, end_color=color, fill_type='solid')
    if value > 0:
        color = green
    if value >= limits[0]:
        color = yellow
    if value > limits[1]:
        color = red

    return PatternFill(start_color=color, end_color=color, fill_type='solid')

def _load_and_store_data():
    workbook = _create_workbook()
    worksheet = workbook.active
    house_listings, land_listings = _get_results()
    calculate_rating(house_listings)
    calculate_rating(land_listings)
    house_listings = sorted(house_listings, key=lambda listing: (
        listing.type.name, listing.rating, listing.ratio))
    land_listings = sorted(land_listings, key=lambda listing: (
        listing.type.name, listing.rating, listing.ratio))

    _write_to_worksheet(worksheet, house_listings)

    worksheet = workbook.worksheets[1]
    
    _write_to_worksheet(worksheet, land_listings)

    workbook.save('immo-results.xlsx')

def _write_to_worksheet(worksheet, listings):
    for row, listing in enumerate(listings, start=2):
        worksheet.cell(row=row, column=1, value=listing.type.name)
        worksheet.cell(row=row, column=2, value=listing.title)
        worksheet.cell(row=row, column=3, value=listing.price)
        worksheet.cell(row=row, column=4, value=listing.living_area)
        worksheet.cell(row=row, column=5, value=listing.land_area)
        worksheet.cell(row=row, column=5,
                       value=listing.ratio).fill = _get_cell_style(listing.type, listing.ratio)
        worksheet.cell(row=row, column=6, value=listing.link)
        worksheet.cell(row=row, column=7, value=listing.rating).fill = _get_cell_style(listing.type, listing.ratio)
        worksheet.cell(row=row, column=8, value=listing.distance)

def _get_results():
    house_listings = []
    land_listings = []
    for platform in ImmoPlatform:
        houses, lands = _get_result_for_platform(platform)
        house_listings += houses
        land_listings += lands
    
    return house_listings, land_listings

def _get_result_for_platform(platform: ImmoPlatform):
    env = os.getenv(platform.value)
    if env != 'active':
        return [], []
    
    if platform == ImmoPlatform.IMMONET:
        return get_immonet_results()
    if platform == ImmoPlatform.IMMOSCOUT:
        return get_immoscout_results()
    if platform == ImmoPlatform.IMMOWELT:
        return get_immowelt_results()

# Load and store data in .xls file
_load_and_store_data()
